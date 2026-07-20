"""
Parser for sheet ``Übersicht`` of third-party funding (Drittmittel) reports.

Focus: non-annual PSP parent elements. Cost-center / annual variants can reuse
helpers later.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from apps.finances.report_import.parsers.base import (
    ParsedContact,
    ParsedCostTypeAmounts,
    ParsedPspParent,
    ParsedReportFile,
    ReportParser,
)
from apps.finances.report_import.suffix_map import COST_TYPE_LABELS, SUFFIX_TO_COST_TYPE

_LABEL_VALUE_RE = re.compile(r'^\s*([^:]+):\s*(.*)\s*$')
_DATE_RE = re.compile(r'(\d{1,2})\.(\d{1,2})\.(\d{4})')
_CHILD_SUFFIX_RE = re.compile(r'^(?P<parent>.+)\.(?P<suffix>[1-9])$')


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_label_value(value) -> tuple[str, str] | None:
    text = _cell_str(value)
    if not text or ':' not in text:
        return None
    match = _LABEL_VALUE_RE.match(text)
    if not match:
        return None
    return match.group(1).strip().lower(), match.group(2).strip()


def _parse_german_date(value) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    match = _DATE_RE.search(text)
    if not match:
        return None
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_period_start(value) -> date | None:
    """Use only the first date in Projektlaufzeit (end may be incomplete/wrong)."""
    return _parse_german_date(value)


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _cell_str(value).replace(' ', '').replace(',', '.')
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_contact(value: str) -> ParsedContact | None:
    text = (value or '').strip()
    if not text:
        return None
    # Excel: "Nachname, Vorname"
    if ',' in text:
        last, first = text.split(',', 1)
        return ParsedContact(last_name=last.strip(), first_name=first.strip())
    parts = text.split()
    if len(parts) >= 2:
        return ParsedContact(last_name=parts[0], first_name=' '.join(parts[1:]))
    return ParsedContact(last_name=text, first_name='')


def is_placeholder_cost_center(code: str) -> bool:
    if not code or not str(code).strip():
        return True
    normalized = str(code).strip()
    if '#' in normalized:
        return True
    if normalized.lower() in {'n/a', 'na', '-', 'none', 'null'}:
        return True
    return False


def split_wbs_code(code: str) -> tuple[str, str | None]:
    """Return (parent_code, suffix_or_None)."""
    text = (code or '').strip()
    match = _CHILD_SUFFIX_RE.match(text)
    if match:
        return match.group('parent'), match.group('suffix')
    return text, None


class UebersichtPspParser(ReportParser):
    """Parse the first logical overview sheet of a Drittmittelbericht workbook."""

    report_kind = 'psp_uebersicht'
    sheet_name = 'Übersicht'

    def parse(self, file_obj, filename: str) -> ParsedReportFile:
        result = ParsedReportFile(filename=filename, report_kind=self.report_kind)
        try:
            data = file_obj.read() if hasattr(file_obj, 'read') else file_obj
            if isinstance(data, bytes):
                wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
            else:
                wb = load_workbook(file_obj, data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001 - surface parse errors in UI
            result.errors.append(f'Could not open workbook: {exc}')
            return result

        try:
            if self.sheet_name not in wb.sheetnames:
                result.errors.append(
                    f'Sheet "{self.sheet_name}" not found. Available: {", ".join(wb.sheetnames)}'
                )
                return result
            ws = wb[self.sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
        finally:
            wb.close()

        parent = self._parse_sheet_rows(rows, filename)
        if parent is None:
            result.errors.append('No parent PSP code found on Übersicht sheet.')
            return result
        result.parents.append(parent)
        result.notes.append(
            'Parsed Übersicht only (other sheets ignored in this import phase).'
        )
        return result

    def _parse_sheet_rows(self, rows: list[list], filename: str) -> ParsedPspParent | None:
        header: dict[str, str] = {}
        planned_end: date | None = None
        cost_center_by_code: dict[str, str] = {}
        last_booking_by_code: dict[str, date | None] = {}
        budget_rows: list[dict] = []

        budget_header_idx = None
        project_cc_header_idx = None

        for r_idx, row in enumerate(rows):
            # Flatten non-empty cells for label detection
            for cell in row:
                parsed = _parse_label_value(cell)
                if not parsed:
                    continue
                label, value = parsed
                if 'förderkennzeichen' in label or 'foerderkennzeichen' in label:
                    header['funding_id'] = value
                elif 'projektdefinition' in label:
                    header['project_definition'] = value
                elif 'ansprechpartner' in label:
                    header['contact'] = value
                elif 'angelegt' in label:
                    header['created_on'] = value
                elif 'projektleiter' in label:
                    header['project_lead'] = value
                elif label == 'status':
                    header['status'] = value
                elif 'projektlaufzeit' in label:
                    header['project_duration'] = value

            # Planned project end sits under label "Geplantes Projektende"
            for c_idx, cell in enumerate(row):
                if _cell_str(cell).lower() == 'geplantes projektende':
                    # value often one row below, same column
                    if r_idx + 1 < len(rows):
                        planned_end = _parse_german_date(rows[r_idx + 1][c_idx])

            cells_lower = [_cell_str(c).lower() for c in row]
            if (
                'projekt' in cells_lower
                and 'kostenstelle' in cells_lower
                and project_cc_header_idx is None
            ):
                project_cc_header_idx = r_idx
                continue
            if (
                'projekt' in cells_lower
                and any('bezeichnung' in c for c in cells_lower)
                and any('budget' in c for c in cells_lower)
            ):
                budget_header_idx = r_idx
                continue

            if project_cc_header_idx is not None and budget_header_idx is None:
                # Table 1 rows until budget header
                code = _cell_str(row[1] if len(row) > 1 else None)  # column B often
                # Find project + cost center by scanning
                code, cc, last_booking = self._extract_project_cc_row(row)
                if code:
                    cost_center_by_code[code] = cc
                    last_booking_by_code[code] = last_booking

            if budget_header_idx is not None and r_idx > budget_header_idx:
                parsed_budget = self._extract_budget_row(row)
                if parsed_budget:
                    budget_rows.append(parsed_budget)

        # Identify parent: prefer codes without suffix that have children, else shortest
        all_codes = set(cost_center_by_code) | {b['code'] for b in budget_rows}
        parents = set()
        children_of: dict[str, list[str]] = {}
        for code in all_codes:
            parent_code, suffix = split_wbs_code(code)
            if suffix:
                parents.add(parent_code)
                children_of.setdefault(parent_code, []).append(suffix)
            else:
                parents.add(code)

        if not parents and all_codes:
            # Fallback: treat codes without suffix as parent
            parents = {c for c in all_codes if split_wbs_code(c)[1] is None}

        if not parents:
            return None

        # One parent per file for current report type
        parent_code = sorted(parents, key=len)[0]
        for p in parents:
            if p in cost_center_by_code or any(b['code'] == p for b in budget_rows):
                parent_code = p
                break

        # Parent designation from budget table
        parent_designation = ''
        for b in budget_rows:
            if b['code'] == parent_code:
                parent_designation = b.get('designation') or ''
                break

        parent_cc = cost_center_by_code.get(parent_code, '')
        # If parent missing in table1, take any child cost center that is not placeholder
        if not parent_cc:
            for code, cc in cost_center_by_code.items():
                p, _s = split_wbs_code(code)
                if p == parent_code and cc and not is_placeholder_cost_center(cc):
                    parent_cc = cc
                    break

        last_years: list[int] = []
        for code, d in last_booking_by_code.items():
            p, _s = split_wbs_code(code)
            if (p == parent_code or code == parent_code) and d:
                last_years.append(d.year)

        cost_types: dict[str, ParsedCostTypeAmounts] = {}
        for b in budget_rows:
            code = b['code']
            p, suffix = split_wbs_code(code)
            if not suffix or p != parent_code:
                continue  # ignore parent row and foreign codes
            if suffix not in SUFFIX_TO_COST_TYPE:
                continue
            cost_types[suffix] = ParsedCostTypeAmounts(
                suffix=suffix,
                label=b.get('designation') or COST_TYPE_LABELS.get(suffix, ''),
                approved_budget=b.get('approved_budget'),
                verfuegt=b.get('verfuegt'),
                obligo=b.get('obligo'),
                personal_obligo=b.get('personal_obligo'),
            )

        period_start = _parse_period_start(header.get('project_duration', ''))
        contact = _parse_contact(header.get('contact', ''))
        report_created_on = _parse_german_date(header.get('created_on', ''))

        warnings = []
        if not cost_types:
            warnings.append('No cost-type child rows (.1–.9) found in budget table.')
        if is_placeholder_cost_center(parent_cc):
            warnings.append(
                f'Cost center from file looks invalid/placeholder: {parent_cc!r}'
            )

        return ParsedPspParent(
            wbs_code=parent_code,
            source_filename=filename,
            third_party_funder_identifier=parent_designation,
            cost_center_code=parent_cc,
            cost_center_is_placeholder=is_placeholder_cost_center(parent_cc),
            period_start=period_start,
            period_end=planned_end,
            contact=contact,
            report_created_on=report_created_on,
            cost_types=cost_types,
            last_booking_years=sorted(set(last_years)),
            warnings=warnings,
        )

    def _extract_project_cc_row(self, row: list) -> tuple[str, str, date | None]:
        """Extract Projekt, Kostenstelle, Letztes Buchungsdatum from a table-1-like row."""
        # Typical layout: B=Projekt, C=Kostenstelle, D=Erstes, F=Letztes
        values = list(row)
        # Pad
        while len(values) < 8:
            values.append(None)

        code = _cell_str(values[1])  # B
        if not code or code.lower() in {'projekt', 'summe', 'summe:'}:
            # try column A
            code = _cell_str(values[0])
        if not code or code.lower() in {'projekt', 'summe', 'summe:', 'psp'}:
            return '', '', None
        # Must look like a WBS code (contains digit)
        if not re.search(r'\d', code):
            return '', '', None

        cc = _cell_str(values[2])  # C
        last_booking = _parse_german_date(values[5])  # F
        if last_booking is None:
            last_booking = _parse_german_date(values[4])
        return code, cc, last_booking

    def _extract_budget_row(self, row: list) -> dict | None:
        values = list(row)
        while len(values) < 12:
            values.append(None)

        code = _cell_str(values[1])  # B
        if not code:
            code = _cell_str(values[0])
        if not code or code.lower().startswith('summe') or code.lower() == 'projekt':
            return None
        if not re.search(r'\d', code):
            return None

        # Skip pure total rows without a real code structure
        designation = _cell_str(values[2])  # C
        return {
            'code': code,
            'designation': designation,
            'approved_budget': _parse_decimal(values[4]),  # E
            'ist_kosten': _parse_decimal(values[5]),  # F (not stored; Verfügt is used)
            'obligo': _parse_decimal(values[6]),  # G
            'personal_obligo': _parse_decimal(values[8]),  # I
            'verfuegt': _parse_decimal(values[10]),  # K
        }
