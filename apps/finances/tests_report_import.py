"""Tests for third-party funding report Übersicht import."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from openpyxl import Workbook

from apps.accounts.models import CustomUser
from apps.accounts.permissions import assign_permissions_to_groups
from apps.finances.models import (
    ContactPerson,
    CostCenter,
    WBSElement,
    WBSElementObligo,
    WBSElementTrueYearlySpending,
    WBSElementYearEstimate,
)
from apps.finances.report_import.parsers.personalkosten import extract_beleg_date_range
from apps.finances.report_import.parsers.uebersicht import UebersichtPspParser
from apps.core.import_tracking import (
    is_report_older_than_last_import,
    record_data_import,
    remaining_scopes_for_hash,
    sha256_bytes,
)
from apps.core.models import DataImportLog
from apps.finances.report_import.service import (
    analyze_uploaded_files,
    apply_import_plan,
    merge_user_decisions,
    normalize_import_scopes,
)


def _build_sample_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Übersicht'
    ws['F3'] = 'Förderkennzeichen:  DFG - TEST-1'
    ws['B4'] = 'Projektdefinition:  DFG - TEST-1'
    ws['B5'] = 'Ansprechpartner:  Muster, Erika'
    ws['B7'] = 'Projektleiter:        Doe, Jane'
    ws['B8'] = 'Projektlaufzeit:     01.01.2026 bis  31.12.2028'
    ws['H8'] = 'Geplantes Projektende'
    ws['H9'] = date(2028, 12, 31)
    ws['B10'] = 'Projekt'
    ws['C10'] = 'Kostenstelle'
    ws['D10'] = 'Erstes Buchungsdatum'
    ws['F10'] = 'Letztes Buchungsdatum'
    ws['B12'] = 'T-100.0001'
    ws['C12'] = '0001/991000'
    ws['F12'] = date(2026, 6, 1)
    ws['B13'] = 'T-100.0001.1'
    ws['C13'] = '0001/991000'
    ws['F13'] = date(2026, 6, 15)
    ws['B14'] = 'T-100.0001.2'
    ws['C14'] = '0001/991000'
    ws['F14'] = date(2027, 1, 10)  # after 2026 for year check
    ws['B17'] = 'Projekt'
    ws['C17'] = 'PSP Bezeichnung'
    ws['E17'] = 'Freigegebenes Budget'
    ws['F17'] = 'Ist-Kosten'
    ws['G17'] = 'Obligo'
    ws['I17'] = 'Personalobligo'
    ws['K17'] = 'Verfügt'
    ws['B18'] = 'T-100.0001'
    ws['C18'] = 'DFG Parent Designation'
    ws['E18'] = 0
    ws['F18'] = 0
    ws['G18'] = 0
    ws['I18'] = 0
    ws['K18'] = 0
    ws['B19'] = 'T-100.0001.1'
    ws['C19'] = 'Sachaufwendungen'
    ws['E19'] = 10000
    ws['F19'] = 1000
    ws['G19'] = 200
    ws['I19'] = 0
    ws['K19'] = 1200
    ws['B20'] = 'T-100.0001.2'
    ws['C20'] = 'Personalaufwendungen'
    ws['E20'] = 50000
    ws['F20'] = 5000
    ws['G20'] = 0
    ws['I20'] = 8000
    ws['K20'] = 13000
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook_with_personalkosten() -> bytes:
    """Sample Übersicht + Personalkosten with Belegdatum range."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Übersicht'
    ws['F3'] = 'Förderkennzeichen:  DFG - TEST-1'
    ws['B4'] = 'Projektdefinition:  DFG - TEST-1'
    ws['B5'] = 'Ansprechpartner:  Muster, Erika'
    ws['B8'] = 'Projektlaufzeit:     01.01.2026 bis  31.12.2028'
    ws['H9'] = date(2028, 12, 31)
    ws['B10'] = 'Projekt'
    ws['C10'] = 'Kostenstelle'
    ws['F10'] = 'Letztes Buchungsdatum'
    ws['B12'] = 'T-200.0001'
    ws['C12'] = '0001/991000'
    ws['F12'] = date(2026, 6, 1)
    ws['B13'] = 'T-200.0001.2'
    ws['C13'] = '0001/991000'
    ws['F13'] = date(2026, 6, 1)
    ws['B17'] = 'Projekt'
    ws['C17'] = 'PSP Bezeichnung'
    ws['E17'] = 'Freigegebenes Budget'
    ws['F17'] = 'Ist-Kosten'
    ws['G17'] = 'Obligo'
    ws['I17'] = 'Personalobligo'
    ws['K17'] = 'Verfügt'
    ws['B18'] = 'T-200.0001'
    ws['C18'] = 'Parent'
    ws['E18'] = 0
    ws['F18'] = 0
    ws['G18'] = 0
    ws['I18'] = 0
    ws['K18'] = 0
    ws['B19'] = 'T-200.0001.2'
    ws['C19'] = 'Personal'
    ws['E19'] = 10000
    ws['F19'] = 1000
    ws['G19'] = 0
    ws['I19'] = 500
    ws['K19'] = 1200

    pk = wb.create_sheet('Personalkosten')
    # Row 2 headers (index 1): B=PSP, C=Kostenart, G=Belegdatum, H=Personalnummer, I=Personalkosten
    pk['B2'] = 'PSP'
    pk['C2'] = 'Kostenart'
    pk['G2'] = 'Belegdatum'
    pk['H2'] = 'Personalnummer'
    pk['I2'] = 'Personalkosten'
    pk['B3'] = 'T-200.0001.2'
    pk['C3'] = '60003000'
    pk['G3'] = date(2026, 1, 15)
    pk['H3'] = '50001001'
    pk['I3'] = 1000
    pk['B4'] = 'T-200.0001.2'
    pk['C4'] = '60003000'
    pk['G4'] = date(2026, 3, 28)
    pk['H4'] = '50001002'
    pk['I4'] = 2000
    # Other cost type with earlier date — still counts for coverage range
    pk['B5'] = 'T-200.0001.2'
    pk['C5'] = '60001000'
    pk['G5'] = date(2025, 12, 1)
    pk['H5'] = '50001003'
    pk['I5'] = 50

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_sap_personalkosten_only_workbook():
    """Real SAP column order including Buchungstext + Buchungsdatum + Personalkosten."""
    from openpyxl import Workbook

    wb = Workbook()
    # Minimal Übersicht so detect_and_parse does not fail hard
    ws = wb.active
    ws.title = 'Übersicht'
    ws['B12'] = 'S-100.0001'
    ws['C12'] = '0001/991000'

    pk = wb.create_sheet('Personalkosten')
    pk['B2'] = 'PSP'
    pk['C2'] = 'Kostenart'
    pk['D2'] = 'KOA Bezeichnung'
    pk['E2'] = 'Buchungstext'
    pk['F2'] = 'Buchungs-\ndatum'
    pk['G2'] = 'Beleg-\ndatum'
    pk['H2'] = 'Personal-\nnummer'
    pk['I2'] = 'Personalkosten'
    pk['B3'] = 'S-100.0001.2'
    pk['C3'] = '60003000'
    pk['D3'] = 'Lohn/Gehalt DA 03'
    pk['E3'] = 'MUSTERSCHMIDT ANNA - Für-Periode 06 - 2026'
    pk['F3'] = '30.06.2026'
    pk['G3'] = '22.06.2026'
    pk['H3'] = '50999001'
    pk['I3'] = 1234.56
    pk['B4'] = 'S-100.0001.2'
    pk['C4'] = '60003500'
    pk['E4'] = 'OTHER PERSON - Für-Periode 06 - 2026'
    pk['F4'] = '30.06.2026'
    pk['H4'] = '50999002'
    pk['I4'] = 9999.00  # DA 35 base salary — included
    pk['B5'] = 'S-100.0001.2'
    pk['C5'] = '61003500'
    pk['E5'] = 'OTHER PERSON - social'
    pk['F5'] = '30.06.2026'
    pk['H5'] = '50999002'
    pk['I5'] = 500.00  # Sozialabgaben — ignored

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class SapPersonalkostenLayoutTests(TestCase):
    def test_parses_sap_layout_with_amount_column(self):
        from apps.finances.report_import.parsers.personalkosten import (
            parse_personalkosten_sheet,
        )

        data = _build_sap_personalkosten_only_workbook()
        entries = parse_personalkosten_sheet(data, 'sap.xlsx')
        by_pn = {e.personalnummer: e for e in entries}
        self.assertEqual(set(by_pn), {'50999001', '50999002'})
        self.assertEqual(by_pn['50999001'].personalkosten, Decimal('1234.56'))
        self.assertEqual(by_pn['50999001'].buchungsdatum, date(2026, 6, 30))
        self.assertIn('MUSTERSCHMIDT', by_pn['50999001'].buchungstext)
        # DA 35 included; social contribution not used instead
        self.assertEqual(by_pn['50999002'].kostenart, '60003500')
        self.assertEqual(by_pn['50999002'].personalkosten, Decimal('9999.00'))


class BelegDateRangeTests(TestCase):
    def test_extract_beleg_from_to(self):
        data = _build_workbook_with_personalkosten()
        beleg_from, beleg_to = extract_beleg_date_range(data)
        self.assertEqual(beleg_from, date(2025, 12, 1))
        self.assertEqual(beleg_to, date(2026, 3, 28))

    def test_analyze_stores_beleg_range_on_meta(self):
        data = _build_workbook_with_personalkosten()
        upload = SimpleUploadedFile('with-pk.xlsx', data)
        plan = analyze_uploaded_files([upload], import_year=2026)
        meta = plan['upload_meta'][0]
        self.assertEqual(meta.get('beleg_from'), '2025-12-01')
        self.assertEqual(meta.get('beleg_to'), '2026-03-28')


class UebersichtParserTests(TestCase):
    def test_parses_parent_and_cost_types(self):
        data = _build_sample_workbook()
        result = UebersichtPspParser().parse(data, 'sample.xlsx')
        self.assertFalse(result.errors)
        self.assertEqual(len(result.parents), 1)
        parent = result.parents[0]
        self.assertEqual(parent.wbs_code, 'T-100.0001')
        self.assertEqual(parent.third_party_funder_identifier, 'DFG Parent Designation')
        self.assertEqual(parent.cost_center_code, '0001/991000')
        self.assertEqual(parent.period_start, date(2026, 1, 1))
        self.assertEqual(parent.period_end, date(2028, 12, 31))
        self.assertEqual(parent.contact.last_name, 'Muster')
        self.assertEqual(parent.contact.first_name, 'Erika')
        self.assertIn('1', parent.cost_types)
        self.assertIn('2', parent.cost_types)
        self.assertEqual(parent.cost_types['1'].approved_budget, Decimal('10000'))
        self.assertEqual(parent.cost_types['1'].verfuegt, Decimal('1200'))
        self.assertEqual(parent.cost_types['2'].personal_obligo, Decimal('8000'))
        self.assertIn(2027, parent.last_booking_years)


class CostCenterLookupTests(TestCase):
    def test_matches_short_code_when_file_has_prefix(self):
        from apps.finances.report_import.service import find_cost_center, get_or_create_cost_center

        existing = CostCenter.objects.create(cost_center='991000')
        match = find_cost_center('0001/991000')
        self.assertEqual(match.pk, existing.pk)

        cc, created = get_or_create_cost_center('0001/991000')
        self.assertFalse(created)
        self.assertEqual(cc.pk, existing.pk)
        self.assertEqual(CostCenter.objects.filter(cost_center__contains='991000').count(), 1)

    def test_matches_prefixed_db_code_when_file_has_short(self):
        from apps.finances.report_import.service import find_cost_center

        existing = CostCenter.objects.create(cost_center='0001/991000')
        match = find_cost_center('991000')
        self.assertEqual(match.pk, existing.pk)

    def test_analyze_reuses_short_cost_center(self):
        CostCenter.objects.create(cost_center='991000')
        data = _build_sample_workbook()  # file uses 0001/991000
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files([upload], import_year=2026)
        parent = plan['parents'][0]
        self.assertTrue(parent['cost_center']['exists_in_db'])
        self.assertFalse(parent['cost_center']['will_create'])
        self.assertEqual(parent['cost_center']['matched_code'], '991000')
        self.assertEqual(parent['cost_center']['matched_via'], 'prefix_stripped')


class ReportImportServiceTests(TestCase):
    def test_normalize_import_scopes(self):
        defaults = normalize_import_scopes(None)
        self.assertTrue(defaults['psp'])
        self.assertTrue(defaults['personnel'])
        self.assertFalse(defaults['orders'])

        from_form = normalize_import_scopes({'scope_psp': 'on'})
        self.assertTrue(from_form['psp'])
        self.assertFalse(from_form['personnel'])

        personnel_only = normalize_import_scopes({'scope_personnel': 'on'})
        self.assertFalse(personnel_only['psp'])
        self.assertTrue(personnel_only['personnel'])

    def test_analyze_and_commit_create_flow(self):
        data = _build_sample_workbook()
        upload = SimpleUploadedFile(
            'sample.xlsx',
            data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        plan = analyze_uploaded_files([upload], import_year=2026)
        self.assertEqual(len(plan['parents']), 1)
        self.assertTrue(plan['import_scopes']['psp'])
        parent = plan['parents'][0]
        self.assertEqual(parent['action'], 'create')
        self.assertTrue(parent['needs_title'])
        self.assertTrue(parent['year_plausibility_warning'])

        post = {
            'title__T-100.0001': 'Test Project Title',
            'confirm_import_year': 'on',
        }
        plan, errors = merge_user_decisions(plan, post)
        self.assertEqual(errors, [])
        summary = apply_import_plan(plan)
        self.assertEqual(summary['psp_created'], 1)

        wbs = WBSElement.objects.get(wbs_code='T-100.0001')
        self.assertEqual(wbs.title, 'Test Project Title')
        self.assertEqual(wbs.third_party_funder_identifier, 'DFG Parent Designation')
        self.assertFalse(wbs.subject_to_annual_recurrence)
        self.assertTrue(wbs.has_material_costs)
        self.assertTrue(wbs.has_personnel_costs)
        self.assertEqual(wbs.cost_center.cost_center, '0001/991000')
        self.assertEqual(wbs.contact_person.last_name, 'Muster')

        # Non-annual: single lifetime plan (technical year = project start year 2026)
        self.assertEqual(wbs.year_estimates.count(), 1)
        ye = wbs.year_estimates.get()
        self.assertEqual(ye.year, 2026)
        self.assertEqual(ye.material_costs, Decimal('10000'))
        self.assertEqual(ye.personnel_costs, Decimal('50000'))

    def test_psp_scope_only_skips_personnel(self):
        data = _build_sample_workbook()
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files(
            [upload],
            import_year=2026,
            import_scopes={'psp': True, 'personnel': False},
        )
        self.assertEqual(plan['personnel_checks'], [])
        self.assertFalse(plan['requires_personnel_resolution'])
        plan, errors = merge_user_decisions(plan, {
            'title__T-100.0001': 'Only PSP',
            'confirm_import_year': 'on',
        })
        self.assertEqual(errors, [])
        summary = apply_import_plan(plan)
        self.assertEqual(summary['psp_created'], 1)
        self.assertEqual(summary['personnel_notes'], [])

    def test_personnel_scope_only_does_not_create_psp(self):
        data = _build_sample_workbook()
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files(
            [upload],
            import_year=2026,
            import_scopes={'psp': False, 'personnel': True},
        )
        self.assertFalse(plan['import_scopes']['psp'])
        self.assertFalse(plan['requires_year_confirmation'])
        # No title required when PSP scope off
        plan, errors = merge_user_decisions(plan, {})
        self.assertEqual(errors, [])
        summary = apply_import_plan(plan)
        self.assertEqual(summary['psp_created'], 0)
        self.assertFalse(WBSElement.objects.filter(wbs_code='T-100.0001').exists())

    def test_same_file_hash_blocks_already_imported_scopes(self):
        data = _build_sample_workbook()
        file_hash = sha256_bytes(data)
        record_data_import(
            kind=DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            uploaded_by=None,
            original_filename='prev.xlsx',
            file_sha256=file_hash,
            file_size=len(data),
            report_created_on=date(2026, 1, 15),
            status=DataImportLog.Status.COMPLETED,
            summary='scopes=psp,personnel; test prior full import',
        )
        already, remaining, requested = remaining_scopes_for_hash(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            file_hash,
            {'psp': True, 'personnel': True},
        )
        self.assertEqual(already, {'psp', 'personnel'})
        self.assertEqual(remaining, set())

        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files([upload], import_year=2026)
        self.assertTrue(plan['has_duplicate_files'])
        self.assertTrue(plan['upload_meta'][0]['is_duplicate'])

    def test_partial_scope_reimport_allowed_for_same_hash(self):
        data = _build_sample_workbook()
        file_hash = sha256_bytes(data)
        record_data_import(
            kind=DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            uploaded_by=None,
            original_filename='prev.xlsx',
            file_sha256=file_hash,
            file_size=len(data),
            report_created_on=date(2026, 1, 15),
            status=DataImportLog.Status.COMPLETED,
            summary='scopes=psp; only psp before',
        )
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files(
            [upload],
            import_year=2026,
            import_scopes={'psp': True, 'personnel': True},
        )
        self.assertFalse(plan['has_duplicate_files'])
        self.assertEqual(plan['upload_meta'][0]['scopes_already_imported'], ['psp'])
        self.assertIn('personnel', plan['upload_meta'][0]['scopes_remaining'])

    def test_older_report_creation_date_is_blocked(self):
        record_data_import(
            kind=DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            uploaded_by=None,
            original_filename='newer.xlsx',
            file_sha256='a' * 64,
            report_created_on=date(2026, 6, 1),
            status=DataImportLog.Status.COMPLETED,
            summary='scopes=psp,personnel',
        )
        is_older, upload_date, prior, prior_date = is_report_older_than_last_import(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            report_created_on=date(2026, 3, 1),
        )
        self.assertTrue(is_older)
        self.assertEqual(upload_date, date(2026, 3, 1))
        self.assertEqual(prior_date, date(2026, 6, 1))

        is_ok, _, _, _ = is_report_older_than_last_import(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            report_created_on=date(2026, 6, 1),
        )
        self.assertFalse(is_ok)

        is_newer, _, _, _ = is_report_older_than_last_import(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            report_created_on=date(2026, 7, 1),
        )
        self.assertFalse(is_newer)

    def test_lifetime_plan_overwrites_single_row_not_import_year_key(self):
        """Re-import updates the one lifetime row even if technical year ≠ import year."""
        cc = CostCenter.objects.create(cost_center='0001/991000')
        wbs = WBSElement.objects.create(
            wbs_code='T-100.0001',
            title='Existing',
            cost_center=cc,
            period_start=date(2025, 1, 1),
            period_end=date(2028, 12, 31),
            subject_to_annual_recurrence=False,
            has_material_costs=True,
        )
        WBSElementYearEstimate.objects.create(
            wbs_element=wbs,
            year=2025,
            material_costs=Decimal('1.00'),
            personnel_costs=Decimal('2.00'),
        )
        data = _build_sample_workbook()
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files([upload], import_year=2026)
        plan, errors = merge_user_decisions(plan, {
            'confirm_import_year': 'on',
            'update_existing_snapshots': 'on',
        })
        self.assertEqual(errors, [])
        apply_import_plan(plan)
        self.assertEqual(wbs.year_estimates.count(), 1)
        ye = wbs.year_estimates.get()
        self.assertEqual(ye.year, 2025)  # existing technical key kept
        self.assertEqual(ye.material_costs, Decimal('10000'))
        self.assertEqual(ye.personnel_costs, Decimal('50000'))

        true = WBSElementTrueYearlySpending.objects.get(
            wbs_element=wbs, date_of_update=date.today()
        )
        self.assertEqual(true.material_costs, Decimal('1200'))
        self.assertEqual(true.personnel_costs, Decimal('13000'))

        obligo = WBSElementObligo.objects.get(
            wbs_element=wbs, date_of_update=date.today()
        )
        self.assertEqual(obligo.material_costs, Decimal('200'))
        self.assertEqual(obligo.personal, Decimal('8000'))

    def test_snapshot_conflict_requires_update_flag(self):
        CostCenter.objects.create(cost_center='0001/991000')
        wbs = WBSElement.objects.create(
            wbs_code='T-100.0001',
            title='Existing',
            cost_center=CostCenter.objects.get(cost_center='0001/991000'),
        )
        WBSElementTrueYearlySpending.objects.create(
            wbs_element=wbs,
            date_of_update=date.today(),
            material_costs=1,
        )
        data = _build_sample_workbook()
        upload = SimpleUploadedFile('sample.xlsx', data)
        plan = analyze_uploaded_files([upload], import_year=2026)
        self.assertTrue(plan['requires_snapshot_update_option'])
        plan, errors = merge_user_decisions(plan, {'confirm_import_year': 'on'})
        self.assertTrue(errors)
        plan, errors = merge_user_decisions(plan, {
            'confirm_import_year': 'on',
            'update_existing_snapshots': 'on',
        })
        self.assertEqual(errors, [])
        apply_import_plan(plan)
        true = WBSElementTrueYearlySpending.objects.get(
            wbs_element=wbs, date_of_update=date.today()
        )
        self.assertEqual(true.material_costs, Decimal('1200'))


class ReportImportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        assign_permissions_to_groups()

    def setUp(self):
        self.user = CustomUser.objects.create_user('importer', password='test')
        self.user.password_changed = True
        self.user.save(update_fields=['password_changed'])
        perm = Permission.objects.filter(
            codename='import_third_party_funding_report'
        ).first()
        if perm:
            self.user.user_permissions.add(perm)
        self.client = Client()
        self.client.login(username='importer', password='test')

    def test_upload_page_requires_permission(self):
        other = CustomUser.objects.create_user('nope', password='test')
        other.password_changed = True
        other.save(update_fields=['password_changed'])
        c = Client()
        c.login(username='nope', password='test')
        response = c.get('/finances/import/third-party-funding/')
        self.assertEqual(response.status_code, 403)

    def test_upload_page_ok_for_permission(self):
        response = self.client.get('/finances/import/third-party-funding/')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Reference year')
