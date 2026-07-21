"""Tests for data import logging and Excel metadata helpers."""

from datetime import date, datetime, timezone as dt_timezone
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook

from apps.core.import_tracking import (
    extract_xlsx_document_timestamps,
    find_completed_import_by_hash,
    parse_scopes_from_import_summary,
    record_data_import,
    remaining_scopes_for_hash,
    sha256_bytes,
)
from apps.core.models import DataImportLog
from apps.finances.report_import.service import (
    SCOPE_PERSONNEL,
    SCOPE_PSP,
    analyze_uploaded_files,
    apply_import_plan,
)


User = get_user_model()


def _minimal_xlsx_with_core_props(created_iso: str) -> bytes:
    """Build a minimal xlsx and inject core properties with a known created date."""
    buf = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Übersicht'
    ws['B4'] = 'Projektdefinition:  Test'
    ws['B5'] = 'Ansprechpartner:  Doe, Jane'
    ws['F5'] = 'angelegt am:               15.03.2025'
    ws['B8'] = 'Projektlaufzeit:     01.01.2026 bis  31.12.2027'
    ws['H8'] = 'Geplantes Projektende'
    ws['H9'] = date(2027, 12, 31)
    ws['B10'] = 'Projekt'
    ws['C10'] = 'Kostenstelle'
    ws['B12'] = 'Z-100.0001'
    ws['C12'] = '0001/991000'
    ws['B17'] = 'Projekt'
    ws['C17'] = 'PSP Bezeichnung'
    ws['E17'] = 'Freigegebenes Budget'
    ws['F17'] = 'Ist-Kosten'
    ws['G17'] = 'Obligo'
    ws['I17'] = 'Personalobligo'
    ws['K17'] = 'Verfügt'
    ws['B18'] = 'Z-100.0001'
    ws['C18'] = 'Test Project Designation'
    ws['E18'] = 0
    ws['K18'] = 0
    ws['B19'] = 'Z-100.0001.1'
    ws['C19'] = 'Sachaufwendungen'
    ws['E19'] = 1000
    ws['K19'] = 100
    wb.save(buf)
    raw = buf.getvalue()

    # Inject / replace core.xml with known timestamps
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties '
        'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created_iso}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created_iso}</dcterms:modified>'
        '</cp:coreProperties>'
    )
    out = BytesIO()
    with ZipFile(BytesIO(raw), 'r') as zin, ZipFile(out, 'w', ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'docProps/core.xml':
                data = core_xml.encode('utf-8')
            zout.writestr(item, data)
        if 'docProps/core.xml' not in zin.namelist():
            zout.writestr('docProps/core.xml', core_xml.encode('utf-8'))
    return out.getvalue()


class ImportTrackingHelpersTests(TestCase):
    def test_sha256_and_ooxml_created(self):
        content = _minimal_xlsx_with_core_props('2026-07-16T12:05:28Z')
        digest = sha256_bytes(content)
        self.assertEqual(len(digest), 64)
        created, modified = extract_xlsx_document_timestamps(content)
        self.assertIsNotNone(created)
        self.assertEqual(created.year, 2026)
        self.assertEqual(created.month, 7)
        self.assertEqual(created.day, 16)

    def test_duplicate_detection_blocks_same_scopes_again(self):
        user = User.objects.create_user('importer', password='test')
        content = _minimal_xlsx_with_core_props('2026-07-16T12:05:28Z')
        digest = sha256_bytes(content)
        record_data_import(
            kind=DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            uploaded_by=user,
            original_filename='report.xlsx',
            file_sha256=digest,
            file_size=len(content),
            status=DataImportLog.Status.COMPLETED,
            summary='scopes=psp,personnel; year=2026',
        )
        prior = find_completed_import_by_hash(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT, digest
        )
        self.assertIsNotNone(prior)

        upload = SimpleUploadedFile('report.xlsx', content)
        plan = analyze_uploaded_files(
            [upload],
            import_year=2026,
            import_scopes={SCOPE_PSP: True, SCOPE_PERSONNEL: True},
        )
        self.assertTrue(plan['has_duplicate_files'])
        self.assertTrue(plan['has_blocking_errors'])

    def test_same_file_allowed_for_remaining_scope(self):
        user = User.objects.create_user('importer-partial', password='test')
        content = _minimal_xlsx_with_core_props('2026-07-16T12:05:28Z')
        digest = sha256_bytes(content)
        record_data_import(
            kind=DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            uploaded_by=user,
            original_filename='report.xlsx',
            file_sha256=digest,
            file_size=len(content),
            status=DataImportLog.Status.COMPLETED,
            summary='scopes=psp; year=2026',
        )
        already, remaining, requested = remaining_scopes_for_hash(
            DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT,
            digest,
            {SCOPE_PSP: True, SCOPE_PERSONNEL: True},
        )
        self.assertEqual(already, {SCOPE_PSP})
        self.assertEqual(remaining, {SCOPE_PERSONNEL})

        upload = SimpleUploadedFile('report.xlsx', content)
        plan = analyze_uploaded_files(
            [upload],
            import_year=2026,
            import_scopes={SCOPE_PSP: True, SCOPE_PERSONNEL: True},
        )
        self.assertFalse(plan['has_duplicate_files'])
        self.assertFalse(plan['has_blocking_errors'])
        meta = plan['upload_meta'][0]
        self.assertEqual(meta['scopes_already_imported'], ['psp'])
        self.assertEqual(meta['scopes_remaining'], ['personnel'])

    def test_parse_scopes_from_summary(self):
        self.assertEqual(
            parse_scopes_from_import_summary('scopes=psp; year=2026'),
            {'psp'},
        )
        self.assertEqual(
            parse_scopes_from_import_summary('legacy without scopes'),
            {'psp', 'personnel'},
        )

    def test_successful_import_writes_log(self):
        user = User.objects.create_user('importer2', password='test')
        content = _minimal_xlsx_with_core_props('2026-07-16T12:05:28Z')
        upload = SimpleUploadedFile('Drittmittelbericht Z-100.0001.xlsx', content)
        plan = analyze_uploaded_files([upload], import_year=2026)
        self.assertFalse(plan['has_duplicate_files'])
        # Need title for create
        for parent in plan['parents']:
            parent['needs_title'] = True
            parent['proposed_title'] = 'Test title'
        from apps.finances.report_import.service import merge_user_decisions
        plan, errors = merge_user_decisions(plan, {
            'title__Z-100.0001': 'Test title',
            'confirm_import_year': 'on',
            'update_existing_snapshots': 'on',
        })
        self.assertEqual(errors, [])
        summary = apply_import_plan(plan, uploaded_by=user)
        self.assertEqual(summary['import_logs'], 1)
        log = DataImportLog.objects.get()
        self.assertEqual(log.kind, DataImportLog.Kind.THIRD_PARTY_FUNDING_REPORT)
        self.assertEqual(log.uploaded_by, user)
        self.assertTrue(log.original_filename.endswith('.xlsx'))
        self.assertEqual(len(log.file_sha256), 64)
        self.assertIsNotNone(log.file_created_at)
