"""
Parser for sheet ``Personalkosten`` of third-party funding reports.

Only cost type (Kostenart) 60003000, positive Personalkosten amounts.
Per Personalnummer keep the row with the latest Buchungsdatum
(fallback: Belegdatum). Negative amounts are ignored.

Note: SAP Excel exports must be opened with openpyxl ``read_only=False``.
In read_only mode the last column (Personalkosten) is frequently dropped.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from apps.finances.report_import.parsers.uebersicht import (
    _cell_str,
    _parse_german_date,
    split_wbs_code,
)

# Legacy single-type constant (tests / callers may still reference it).
RELEVANT_KOSTENART = '60003000'

# Base monthly Lohn/Gehalt lines used for personnel matching.
# Examples from real reports:
#   60003000 Lohn/Gehalt DA 03
#   60003500 Lohn/Gehalt DA 35
#   60013000 Lohn/Gehalt DA 13
# Excluded (not base monthly salary):
#   6080xxxx Jahressonderzahlung
#   6100xxxx Sozialabgaben
#   6200xxxx Altersversorgung
#   6481xxxx sonst. Personalaufwand


def is_relevant_gehalt_kostenart(kostenart: str) -> bool:
    """True for base monthly Lohn/Gehalt cost types (not social/pension/bonus)."""
    ka = (kostenart or '').strip()
    if not ka:
        return False
    if ka.startswith('608'):
        return False
    # 6000xxxx / 6001xxxx = Lohn/Gehalt DA variants
    return ka.startswith('6000') or ka.startswith('6001')

# Known SAP export layout (Excel columns B–I → 0-based indices 1–8)
_SAP_FALLBACK_COLUMNS = {
    'psp': 1,
    'kostenart': 2,
    'buchungstext': 4,
    'buchungsdatum': 5,
    'belegdatum': 6,
    'personalnummer': 7,
    'personalkosten': 8,
}


@dataclass
class ParsedPersonalkostenEntry:
    personalnummer: str
    kostenart: str
    personalkosten: Decimal
    belegdatum: date | None
    buchungsdatum: date | None
    buchungstext: str
    psp_code: str
    parent_psp_code: str
    source_filename: str

    @property
    def booking_date(self) -> date | None:
        """Date used for month selection / FA window (Buchungsdatum preferred)."""
        return self.buchungsdatum or self.belegdatum


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(' ', '').replace(',', '.')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _normalize_personalnummer(value) -> str:
    text = _cell_str(value)
    if not text:
        return ''
    # Excel may store as number → "50001094.0"
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text.strip()


def _normalize_kostenart(value) -> str:
    text = _cell_str(value)
    if not text:
        return ''
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    return text.strip()


def parse_name_from_buchungstext(buchungstext: str) -> tuple[str, str]:
    """
    Extract (last_name, first_name) from leading fully-uppercase tokens.

    Last name = first token; remainder of uppercase run = first name.
    """
    text = (buchungstext or '').strip()
    if not text:
        return 'Unknown', 'Unknown'

    parts = text.split()
    upper_parts: list[str] = []
    for part in parts:
        letters = ''.join(ch for ch in part if ch.isalpha())
        if letters and letters == letters.upper():
            upper_parts.append(part)
        else:
            break

    if not upper_parts:
        if len(parts) == 1:
            return parts[0].title(), 'Unknown'
        return parts[0].title(), ' '.join(parts[1:]).title()

    last = upper_parts[0].title()
    first = ' '.join(upper_parts[1:]).title() if len(upper_parts) > 1 else 'Unknown'
    return last, first


def _load_personalkosten_rows(file_bytes: bytes) -> list:
    """
    Load Personalkosten sheet rows.

    Do **not** use openpyxl ``read_only=True`` for these SAP exports — the
    Personalkosten amount column is often omitted in that mode.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    except Exception:  # noqa: BLE001
        return []
    try:
        if 'Personalkosten' not in wb.sheetnames:
            return []
        ws = wb['Personalkosten']
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _detect_personalkosten_layout(rows: list) -> tuple[int, dict]:
    """
    Return (header_row_index, column_map) for a Personalkosten sheet.

    Real SAP layout (column B onwards)::
        PSP | Kostenart | KOA Bezeichnung | Buchungstext |
        Buchungsdatum | Belegdatum | Personalnummer | Personalkosten
    """
    header_idx = None
    col: dict = {}
    for i, row in enumerate(rows[:15]):
        cells = [
            _cell_str(c).lower().replace('\n', ' ').replace('-', ' ')
            for c in row
        ]
        joined = ' '.join(cells)
        has_personal = 'personal' in joined
        has_kostenart = 'kostenart' in joined
        has_amount = 'personalkosten' in joined.replace(' ', '')
        if has_kostenart and has_personal and (has_amount or 'psp' in joined):
            header_idx = i
            for j, c in enumerate(cells):
                c_compact = c.replace(' ', '')
                if 'psp' in c and 'bezeichnung' not in c:
                    col['psp'] = j
                elif c == 'kostenart' or (
                    c.startswith('kostenart')
                    and 'bezeichnung' not in c
                    and 'koa' not in c
                ):
                    col.setdefault('kostenart', j)
                elif 'buchung' in c and 'datum' in c:
                    col['buchungsdatum'] = j
                elif 'buchung' in c and 'text' in c:
                    col['buchungstext'] = j
                elif 'beleg' in c and 'datum' in c:
                    col['belegdatum'] = j
                elif 'personal' in c and 'nummer' in c:
                    col['personalnummer'] = j
                elif 'personalkosten' in c_compact:
                    col['personalkosten'] = j
            break

    if header_idx is None:
        header_idx = 1
        col = dict(_SAP_FALLBACK_COLUMNS)
        if rows and len(rows[0]) > 1 and 'psp' in _cell_str(rows[0][1]).lower():
            header_idx = 0

    required = ('psp', 'kostenart', 'personalnummer', 'personalkosten')
    if not all(k in col for k in required):
        col = dict(_SAP_FALLBACK_COLUMNS)
        header_idx = 1

    return header_idx, col


def extract_beleg_date_range(file_bytes: bytes) -> tuple[date | None, date | None]:
    """
    Heuristic coverage window from Personalkosten Belegdatum / Buchungsdatum.

    Scans **all** data rows with a parseable date (not only Kostenart
    60003000 / latest-per-employee). Returns (min_date, max_date) or (None, None).
    """
    rows = _load_personalkosten_rows(file_bytes)
    if not rows:
        return None, None

    header_idx, col = _detect_personalkosten_layout(rows)
    date_idxs = [
        idx for key in ('belegdatum', 'buchungsdatum')
        if (idx := col.get(key)) is not None
    ]
    if not date_idxs:
        return None, None

    dates: list[date] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        for beleg_idx in date_idxs:
            if beleg_idx >= len(row):
                continue
            d = _parse_german_date(row[beleg_idx])
            if d is not None:
                dates.append(d)

    if not dates:
        return None, None
    return min(dates), max(dates)


def parse_personalkosten_sheet(file_bytes: bytes, filename: str) -> list[ParsedPersonalkostenEntry]:
    """
    Parse Personalkosten sheet; return latest positive base-salary row per Personalnummer.

    Relevant Kostenarten: Lohn/Gehalt (6000xxxx / 6001xxxx), not social, pension,
    or Jahressonderzahlung. “Latest” = newest Buchungsdatum (fallback Belegdatum).
    """
    rows = _load_personalkosten_rows(file_bytes)
    if not rows:
        return []

    header_idx, col = _detect_personalkosten_layout(rows)

    best_by_personal: dict[str, ParsedPersonalkostenEntry] = {}

    for row in rows[header_idx + 1 :]:
        if not row:
            continue

        def get(key, default=None):
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        kostenart = _normalize_kostenart(get('kostenart'))
        if not is_relevant_gehalt_kostenart(kostenart):
            continue

        amount = _parse_decimal(get('personalkosten'))
        if amount is None or amount <= 0:
            continue

        personalnummer = _normalize_personalnummer(get('personalnummer'))
        if not personalnummer:
            continue

        psp_code = _cell_str(get('psp'))
        if not psp_code:
            continue
        parent, _suffix = split_wbs_code(psp_code)

        beleg = None
        if 'belegdatum' in col:
            beleg = _parse_german_date(get('belegdatum'))
        buchung = None
        if 'buchungsdatum' in col:
            buchung = _parse_german_date(get('buchungsdatum'))
        buchungstext = _cell_str(get('buchungstext')) if 'buchungstext' in col else ''

        entry = ParsedPersonalkostenEntry(
            personalnummer=personalnummer,
            kostenart=kostenart,
            personalkosten=amount,
            belegdatum=beleg,
            buchungsdatum=buchung,
            buchungstext=buchungstext,
            psp_code=psp_code,
            parent_psp_code=parent,
            source_filename=filename,
        )

        prev = best_by_personal.get(personalnummer)
        if prev is None:
            best_by_personal[personalnummer] = entry
            continue
        prev_d = prev.booking_date or date.min
        new_d = entry.booking_date or date.min
        if new_d > prev_d:
            best_by_personal[personalnummer] = entry
        elif new_d == prev_d and entry.personalkosten > prev.personalkosten:
            best_by_personal[personalnummer] = entry

    return list(best_by_personal.values())
